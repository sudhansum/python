import openpyxl as op

z = op.Workbook()
print(z.sheetnames)
s = z['Sheet']
s.title = "hello"
z.create_sheet()
z.create_sheet(title="python")
#z.save("C:\\Users\\SUDHANSU\\Desktop\\c.xlsx")
s.cell(row=1,column=2).value ='formyscholars'
s.cell(row=5,column=2).value =20
s.merge_cells("A6:B6")
#s.unmerge_cell("A6:B6")
x = s.cell(row=6,column=1).value ="hello"
#d =op.drawing.image.Image("C:\\Users\\SUDHANSU\\Desktop\\hawk.jpg")
#s.add_image(d,"A1")
f = ((1,2,3),(3,2,8),(4,5,6))
for g in f:
    s.append(g)

s["C11"].value="=SUM(A7:C9)"

z.save("C:\\Users\\SUDHANSU\\Desktop\\b.xlsx")
